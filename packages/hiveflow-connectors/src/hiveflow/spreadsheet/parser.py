"""Detect table regions in Excel workbooks."""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

JOB_KIND = "spreadsheet_engine_parse"
MAX_SAMPLE_ROWS = 25
MIN_DATA_ROWS = 2
MIN_DATA_COLS = 2
MIN_HEADER_LABELS = 2
MIN_HEADER_DENSITY = 0.4

_REPORT_PREAMBLE_RE = re.compile(
    r"(?:\bphone\b|\bfax\b|\bas of\b|\bprinted\b|\bgenerated\b|"
    r"\bconfidential\b|https?://|@)",
    re.IGNORECASE,
)
_LONG_DATETIME_RE = re.compile(
    r"\b(?:monday|tuesday|wednesday|thursday|friday|saturday|sunday)\b.*\b\d{4}\b",
    re.IGNORECASE,
)


def _cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _row_values(sheet: Worksheet, row_idx: int, *, min_col: int, max_col: int) -> list[Any]:
    return [
        _cell_value(sheet.cell(row=row_idx, column=col_idx).value)
        for col_idx in range(min_col, max_col + 1)
    ]


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _row_is_empty(values: list[Any]) -> bool:
    return all(_is_blank(v) for v in values)


def _looks_like_column_header(value: Any) -> bool:
    if _is_blank(value):
        return False
    if isinstance(value, (int, float, bool, datetime, date)):
        return False
    text = str(value).strip()
    if not text or len(text) > 60:
        return False
    if text.endswith(":"):
        return False
    if _REPORT_PREAMBLE_RE.search(text):
        return False
    if _LONG_DATETIME_RE.search(text):
        return False
    if len(text) > 40 and text.count(" ") >= 3:
        return False
    return True


def _header_label_indices(values: list[Any]) -> list[int]:
    return [idx for idx, value in enumerate(values) if _looks_like_column_header(value)]


def _header_band_stats(values: list[Any]) -> tuple[int, int, float]:
    label_indices = _header_label_indices(values)
    if not label_indices:
        return 0, 0, 0.0
    span = label_indices[-1] - label_indices[0] + 1
    return len(label_indices), span, len(label_indices) / span


def _data_supports_headers(
    header_values: list[Any],
    data_rows: list[list[Any]],
    *,
    col_indices: list[int],
) -> bool:
    label_cols = [idx for idx in col_indices if _looks_like_column_header(header_values[idx])]
    if len(label_cols) < MIN_HEADER_LABELS:
        return False
    if not data_rows:
        return True

    supported = 0
    for idx in label_cols:
        if any(idx < len(row) and not _is_blank(row[idx]) for row in data_rows):
            supported += 1
    return supported >= MIN_HEADER_LABELS


def _is_likely_table_header(
    header_values: list[Any],
    data_rows: list[list[Any]],
    *,
    col_indices: list[int],
) -> bool:
    label_count, span, density = _header_band_stats(header_values)
    if label_count < MIN_HEADER_LABELS:
        return False
    if span > 3 and density < MIN_HEADER_DENSITY:
        return False
    return _data_supports_headers(header_values, data_rows, col_indices=col_indices)


def _normalize_header(value: Any, *, index: int) -> str:
    if value is None:
        return f"column_{index + 1}"
    text = str(value).strip()
    if not text:
        return f"column_{index + 1}"
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_").lower()
    return slug or f"column_{index + 1}"


def _sheet_bounds(sheet: Worksheet) -> tuple[int, int, int, int]:
    if sheet.max_row is None or sheet.max_column is None:
        calculate = getattr(sheet, "calculate_dimension", None)
        if callable(calculate):
            calculate(force=True)
    min_row = sheet.min_row or 1
    max_row = sheet.max_row or 1
    min_col = sheet.min_column or 1
    max_col = sheet.max_column or 1
    return min_row, max_row, min_col, max_col


def _trim_data_end(
    sheet: Worksheet,
    *,
    data_start: int,
    data_end: int,
    min_col: int,
    max_col: int,
) -> int:
    end = data_end
    while end >= data_start:
        if not _row_is_empty(_row_values(sheet, end, min_col=min_col, max_col=max_col)):
            return end
        end -= 1
    return data_start - 1


def _scan_data_end(
    sheet: Worksheet,
    *,
    header_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> int:
    data_end = header_row
    blank_streak = 0
    for data_row in range(header_row + 1, max_row + 1):
        values = _row_values(sheet, data_row, min_col=min_col, max_col=max_col)
        if _row_is_empty(values):
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue
        blank_streak = 0
        data_end = data_row
    return data_end


def _row_occupied(row: list[Any], col_indices: list[int]) -> bool:
    return any(idx < len(row) and not _is_blank(row[idx]) for idx in col_indices)


def _should_split_column_groups(groups: list[list[int]], data_rows: list[list[Any]]) -> bool:
    if len(groups) < 2:
        return False
    if not data_rows:
        return True
    overlapping = 0
    disjoint = 0
    for row in data_rows:
        occupied = sum(1 for group in groups if _row_occupied(row, group))
        if occupied >= 2:
            overlapping += 1
        elif occupied == 1:
            disjoint += 1
    return overlapping > disjoint


def _used_column_groups(
    header_values: list[Any],
    data_rows: list[list[Any]],
) -> list[list[int]]:
    used_flags = []
    for idx, header in enumerate(header_values):
        used = not _is_blank(header) or any(
            idx < len(row) and not _is_blank(row[idx]) for row in data_rows
        )
        used_flags.append(used)

    groups: list[list[int]] = []
    current: list[int] = []
    for idx, used in enumerate(used_flags):
        if used:
            current.append(idx)
            continue
        if current:
            groups.append(current)
            current = []
    if current:
        groups.append(current)
    return groups


def _build_region(
    sheet: Worksheet,
    *,
    header_row: int,
    data_end_row: int,
    min_col: int,
    max_col: int,
    excel_table_name: str = "",
    require_header_heuristic: bool = True,
    min_data_rows: int = MIN_DATA_ROWS,
) -> dict[str, Any] | None:
    header_values = _row_values(sheet, header_row, min_col=min_col, max_col=max_col)
    data_start = header_row + 1
    data_end = _trim_data_end(
        sheet,
        data_start=data_start,
        data_end=max(data_end_row, header_row),
        min_col=min_col,
        max_col=max_col,
    )
    data_row_count = max(0, data_end - header_row)
    if data_row_count < min_data_rows:
        return None

    col_count = max_col - min_col + 1
    non_empty_cols = [
        idx
        for idx in range(col_count)
        if not _is_blank(header_values[idx])
        or any(
            not _is_blank(_row_values(sheet, r, min_col=min_col, max_col=max_col)[idx])
            for r in range(data_start, data_end + 1)
        )
    ]
    if len(non_empty_cols) < MIN_DATA_COLS:
        return None

    sample_rows: list[list[Any]] = []
    if data_end >= data_start:
        for sample_row in range(data_start, min(data_end + 1, data_start + MAX_SAMPLE_ROWS)):
            row_vals = _row_values(sheet, sample_row, min_col=min_col, max_col=max_col)
            if _row_is_empty(row_vals):
                continue
            sample_rows.append(row_vals)

    if require_header_heuristic and not _is_likely_table_header(
        header_values,
        sample_rows,
        col_indices=non_empty_cols,
    ):
        return None

    headers = [_normalize_header(header_values[idx], index=idx) for idx in non_empty_cols]
    header_col_offsets = [idx - non_empty_cols[0] for idx in non_empty_cols]
    aligned_sample_rows: list[list[Any]] = []
    for sample_row in sample_rows:
        aligned_sample_rows.append(
            [
                sample_row[offset] if 0 <= offset < len(sample_row) else None
                for offset in header_col_offsets
            ]
        )

    region = {
        "sheet": sheet.title,
        "header_row": header_row,
        "data_start_row": data_start,
        "data_end_row": data_end if data_end >= data_start else header_row,
        "min_col": min_col + non_empty_cols[0],
        "max_col": min_col + non_empty_cols[-1],
        "row_count": data_row_count,
        "column_count": len(headers),
        "headers": headers,
        "header_col_offsets": header_col_offsets,
        "sample_rows": aligned_sample_rows,
    }
    if excel_table_name:
        region["excel_table_name"] = excel_table_name
    return region


def _regions_overlap(left: dict[str, Any], right: dict[str, Any]) -> bool:
    return not (
        int(left["data_end_row"]) < int(right["header_row"])
        or int(right["data_end_row"]) < int(left["header_row"])
        or int(left["max_col"]) < int(right["min_col"])
        or int(right["max_col"]) < int(left["min_col"])
    )


def _excel_table_entries(sheet: Worksheet) -> list[tuple[str, Any]]:
    tables = getattr(sheet, "tables", None)
    if not tables:
        return []
    entries: list[tuple[str, Any]] = []
    for name in list(tables):
        table = tables[name]
        entries.append((str(name), table))
    return entries


def _table_ref(table: Any, *, name: str) -> str:
    if isinstance(table, str):
        return table
    ref = getattr(table, "ref", None)
    if ref:
        return str(ref)
    return name


def _detect_excel_table_regions(sheet: Worksheet) -> list[dict[str, Any]]:
    regions: list[dict[str, Any]] = []
    for name, table in _excel_table_entries(sheet):
        ref = _table_ref(table, name=name)
        try:
            min_col, min_row, max_col, max_row = range_boundaries(ref)
        except (TypeError, ValueError):
            continue
        header_count = 1
        if not isinstance(table, str):
            header_count = int(getattr(table, "headerRowCount", 1) or 1)
        display_name = name
        if not isinstance(table, str):
            display_name = str(getattr(table, "displayName", None) or name)
        data_end = max_row
        region = _build_region(
            sheet,
            header_row=min_row,
            data_end_row=data_end,
            min_col=min_col,
            max_col=max_col,
            excel_table_name=display_name,
            require_header_heuristic=False,
            min_data_rows=0,
        )
        if region is None:
            continue
        if header_count > 1:
            region["data_start_row"] = min_row + header_count
        region["region_kind"] = "excel_table"
        regions.append(region)
    regions.sort(key=lambda item: (item["header_row"], item["min_col"]))
    return regions


def _detect_pivot_table_regions(sheet: Worksheet) -> list[dict[str, Any]]:
    regions: list[dict[str, Any]] = []
    for pivot in list(getattr(sheet, "_pivots", None) or []):
        location = getattr(pivot, "location", None)
        ref = getattr(location, "ref", None) if location is not None else None
        if not ref:
            continue
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(ref))
        except (TypeError, ValueError):
            continue
        first_header = int(getattr(location, "firstHeaderRow", None) or 1)
        first_data = int(getattr(location, "firstDataRow", None) or max(first_header + 1, 2))
        header_row = min_row + max(first_header, 1) - 1
        if header_row < min_row:
            header_row = min_row
        region = _build_region(
            sheet,
            header_row=header_row,
            data_end_row=max_row,
            min_col=min_col,
            max_col=max_col,
            excel_table_name=str(getattr(pivot, "name", None) or "pivot"),
            require_header_heuristic=False,
            min_data_rows=0,
        )
        if region is None:
            continue
        data_start = min_row + max(first_data, 1) - 1
        if data_start > region["header_row"]:
            region["data_start_row"] = data_start
        region["region_kind"] = "pivot"
        region["pivot_name"] = str(getattr(pivot, "name", None) or region["excel_table_name"])
        regions.append(region)
    regions.sort(key=lambda item: (item["header_row"], item["min_col"]))
    return regions


def _detect_inferred_regions(sheet: Worksheet) -> list[dict[str, Any]]:
    min_row, max_row, min_col, max_col = _sheet_bounds(sheet)
    if max_row < min_row or max_col < min_col:
        return []

    regions: list[dict[str, Any]] = []
    row = min_row
    while row <= max_row:
        header_values = _row_values(sheet, row, min_col=min_col, max_col=max_col)
        if _row_is_empty(header_values):
            row += 1
            continue

        band_end = _scan_data_end(
            sheet,
            header_row=row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        )
        data_start = row + 1
        band_rows = [
            _row_values(sheet, data_row, min_col=min_col, max_col=max_col)
            for data_row in range(data_start, band_end + 1)
            if band_end >= data_start
        ]
        groups = _used_column_groups(header_values, band_rows)
        if not groups:
            row += 1
            continue

        split_side_by_side = _should_split_column_groups(groups, band_rows)
        if not split_side_by_side:
            groups = [[idx for group in groups for idx in group]]
        emitted = False
        last_end = row
        for group in groups:
            col_start = min_col + group[0]
            col_end = min_col + group[-1]
            group_end = _scan_data_end(
                sheet,
                header_row=row,
                max_row=band_end,
                min_col=col_start,
                max_col=col_end,
            )
            region = _build_region(
                sheet,
                header_row=row,
                data_end_row=group_end,
                min_col=col_start,
                max_col=col_end,
                min_data_rows=0 if split_side_by_side else MIN_DATA_ROWS,
            )
            if region is None:
                continue
            regions.append(region)
            emitted = True
            last_end = max(last_end, int(region["data_end_row"]))

        if emitted:
            row = last_end + 1
        else:
            row += 1

    return regions


def _detect_regions(sheet: Worksheet) -> list[dict[str, Any]]:
    named_regions = _detect_excel_table_regions(sheet) + _detect_pivot_table_regions(sheet)
    inferred_regions: list[dict[str, Any]] = []
    for region in _detect_inferred_regions(sheet):
        if any(_regions_overlap(region, named) for named in named_regions):
            continue
        region["region_kind"] = "inferred"
        inferred_regions.append(region)
    regions = named_regions + inferred_regions
    regions.sort(key=lambda item: (item["header_row"], item["min_col"], item.get("sheet") or ""))
    return regions


def parse_workbook(
    path: str | Path,
    *,
    filename: str = "",
    sheet_names: list[str] | None = None,
) -> dict[str, Any]:
    """Parse an Excel workbook into table candidates."""
    workbook_path = Path(path)
    # Random cell access in _detect_regions is incompatible with read_only mode
    # (dimensions are unset and per-cell lookups are very slow).
    workbook = load_workbook(workbook_path, data_only=True)
    all_sheet_names = list(workbook.sheetnames)
    selected = [name for name in (sheet_names or []) if str(name).strip()]
    selected_set = set(selected) if selected else None
    tables: list[dict[str, Any]] = []
    table_index = 0
    table_counts: dict[str, int] = {name: 0 for name in all_sheet_names}
    for sheet_name in all_sheet_names:
        if selected_set is not None and sheet_name not in selected_set:
            continue
        sheet = workbook[sheet_name]
        for region in _detect_regions(sheet):
            table_id = f"t{table_index}"
            table_index += 1
            tables.append({"table_id": table_id, **region})
            table_counts[sheet_name] = table_counts.get(sheet_name, 0) + 1
    workbook.close()
    sheets = [
        {"name": name, "table_count": int(table_counts.get(name, 0))}
        for name in all_sheet_names
    ]
    return {
        "kind": JOB_KIND,
        "filename": filename or workbook_path.name,
        "sheet_count": len(all_sheet_names),
        "sheet_names": all_sheet_names,
        "sheets": sheets,
        "selected_sheets": selected or all_sheet_names,
        "table_count": len(tables),
        "tables": tables,
    }
