"""Tests for spreadsheet workbook parsing."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from meshflow.spreadsheet.parser import parse_workbook
from meshflow.spreadsheet.preview import MAX_PREVIEW_ROWS, extract_table_preview
from meshflow.spreadsheet.profiler import profile_tables
from meshflow.spreadsheet.interpret import interpret_tables


def test_extract_table_preview_limits_rows(tmp_path: Path) -> None:
    path = tmp_path / "preview.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"
    ws.append(["Item", "Price"])
    for idx in range(150):
        ws.append([f"I{idx}", idx + 1])
    wb.save(path)

    payload = parse_workbook(path)
    table = payload["tables"][0]
    preview = extract_table_preview(
        path,
        sheet=table["sheet"],
        data_start_row=table["data_start_row"],
        data_end_row=table["data_end_row"],
        min_col=table["min_col"],
        max_col=table["max_col"],
        headers=table["headers"],
        max_rows=MAX_PREVIEW_ROWS,
    )
    assert preview["preview_row_count"] == MAX_PREVIEW_ROWS
    assert preview["row_count"] == 150
    assert preview["truncated"] is True
    assert preview["rows"][0] == ["I0", 1]


def test_load_table_preview_reads_upload(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("MESHFLOW_DATA_DIR", str(tmp_path))

    from meshflow.spreadsheet.jobs import create_job, load_table_preview, run_parse, store_upload

    path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws.append(["Customer ID", "Company"])
    ws.append(["C1", "Acme"])
    ws.append(["C2", "Beta"])
    wb.save(path)

    job = create_job(filename="sample.xlsx", username="poc")
    store_upload(job["job_id"], filename="sample.xlsx", body=path.read_bytes())
    run_parse(job["job_id"])

    preview = load_table_preview(job["job_id"], "t0")
    assert preview is not None
    assert preview["headers"] == ["customer_id", "company"]
    assert preview["preview_row_count"] == 2
    assert preview["rows"] == [["C1", "Acme"], ["C2", "Beta"]]


def test_load_table_preview_keeps_all_null_source_columns(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("MESHFLOW_DATA_DIR", str(tmp_path))

    import json

    from meshflow.spreadsheet.jobs import create_job, load_table_preview, run_parse, store_upload
    from meshflow.storage.paths import prefix_path, spreadsheet_engine_job_table_key

    path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws.append(["Customer ID", "Notes", "Company"])
    ws.append(["C1", None, "Acme"])
    ws.append(["C2", None, "Beta"])
    wb.save(path)

    job = create_job(filename="sample.xlsx", username="poc")
    store_upload(job["job_id"], filename="sample.xlsx", body=path.read_bytes())
    run_parse(job["job_id"])
    table_path = prefix_path(tmp_path, spreadsheet_engine_job_table_key(job["job_id"], "t0"))
    table_path.parent.mkdir(parents=True, exist_ok=True)
    table_path.write_text(
        json.dumps(
            {
                "table_id": "t0",
                "schema": [
                    {"name": "customer_id", "type": "string"},
                    {"name": "company", "type": "string"},
                ],
            }
        ),
        encoding="utf-8",
    )

    preview = load_table_preview(job["job_id"], "t0")
    assert preview is not None
    assert preview["headers"] == ["customer_id", "notes", "company"]
    assert preview["rows"] == [["C1", None, "Acme"], ["C2", None, "Beta"]]


def test_parse_workbook_detects_table_after_report_preamble(tmp_path: Path) -> None:
    path = tmp_path / "price_list.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["List Price Sheet as of 07/01/27"])
    ws.append(["Phone: +1 425 555 0100", None, None, None, None, None, None, None, None, None, None, None, None, "Page", 1])
    ws.append([None] * 7)
    ws.append(["CRONUS International Ltd."])
    ws.append([None] * 7)
    ws.append(["All Customers"])
    ws.append([None] * 7)
    ws.append([None] * 7)
    ws.append(["No.", "Description", "Unit of Measure Code", "Unit Price"])
    ws.append(["1896-S", "ATHENS Desk", None, None])
    ws.append([None, None, "PCS", 1000.8])
    ws.append(["1900-S", "PARIS Guest Chair, black", None, None])
    ws.append([None, None, "PCS", 192.8])
    wb.save(path)

    payload = parse_workbook(path, filename="price_list.xlsx")
    assert payload["table_count"] == 1
    table = payload["tables"][0]
    assert table["headers"][:2] == ["no", "description"]
    assert table["header_row"] == 9
    assert table["row_count"] >= 4


def _write_side_by_side_metadata_sheet(path: Path, *, named_tables: bool) -> None:
    from openpyxl.worksheet.table import Table

    wb = Workbook()
    ws = wb.active
    ws.title = "Aggregated Metadata"
    ws.append(
        [
            "Report Property",
            "Report Property Value",
            None,
            "Request Property",
            "Request Property Value",
            None,
            "Request Page Option",
            "Request Page Option Value",
            None,
            "Filter",
            "Filter Value",
        ]
    )
    ws.append(
        [
            "Extension ID",
            "abc",
            None,
            "Tenant Id",
            "t1",
            None,
            None,
            None,
            None,
            "Sales Line::Document Type",
            "1",
        ]
    )
    ws.append(
        [
            "Object Name",
            "Customer - Order Summary",
            None,
            "Company name",
            "CRONUS USA, Inc.",
            None,
            None,
            None,
            None,
            "Sales Line::Outstanding Quantity",
            "<>0",
        ]
    )
    ws.append(
        [
            None,
            None,
            None,
            "User name",
            "Operator",
            None,
            None,
            None,
            None,
            None,
            None,
        ]
    )
    if named_tables:
        ws.add_table(Table(displayName="ReportMetadataValues", ref="A1:B3"))
        ws.add_table(Table(displayName="ReportRequestValues", ref="D1:E4"))
        ws.add_table(Table(displayName="ReportRequestPageValues", ref="G1:H4"))
        ws.add_table(Table(displayName="ReportFilterValues", ref="J1:K3"))
    wb.save(path)


def test_parse_workbook_splits_excel_list_objects(tmp_path: Path) -> None:
    path = tmp_path / "metadata.xlsx"
    _write_side_by_side_metadata_sheet(path, named_tables=True)

    payload = parse_workbook(path, filename="metadata.xlsx")
    meta = [table for table in payload["tables"] if table["sheet"] == "Aggregated Metadata"]
    assert len(meta) == 4
    assert [table["headers"] for table in meta] == [
        ["report_property", "report_property_value"],
        ["request_property", "request_property_value"],
        ["request_page_option", "request_page_option_value"],
        ["filter", "filter_value"],
    ]
    assert [table.get("excel_table_name") for table in meta] == [
        "ReportMetadataValues",
        "ReportRequestValues",
        "ReportRequestPageValues",
        "ReportFilterValues",
    ]
    assert meta[0]["min_col"] == 1
    assert meta[0]["max_col"] == 2
    assert meta[3]["min_col"] == 10
    assert meta[2]["row_count"] == 0


def test_parse_workbook_splits_side_by_side_tables_without_list_objects(
    tmp_path: Path,
) -> None:
    path = tmp_path / "metadata.xlsx"
    _write_side_by_side_metadata_sheet(path, named_tables=False)

    payload = parse_workbook(path, filename="metadata.xlsx")
    assert payload["table_count"] == 4
    headers = [tuple(table["headers"]) for table in payload["tables"]]
    assert headers == [
        ("report_property", "report_property_value"),
        ("request_property", "request_property_value"),
        ("request_page_option", "request_page_option_value"),
        ("filter", "filter_value"),
    ]
    assert payload["tables"][0]["sample_rows"][0][:2] == ["Extension ID", "abc"]
    assert payload["tables"][3]["sample_rows"][0] == [
        "Sales Line::Document Type",
        "1",
    ]


def _write_pivot_workbook(path: Path) -> None:
    from openpyxl.pivot.cache import CacheDefinition, CacheField, CacheSource, WorksheetSource
    from openpyxl.pivot.table import Location, TableDefinition
    from openpyxl.worksheet.table import Table

    wb = Workbook()
    source = wb.active
    source.title = "Sales"
    source.append(["Region", "Quarter", "Amount"])
    source.append(["East", "Q1", 10])
    source.append(["East", "Q2", 20])
    source.append(["West", "Q1", 30])
    source.append(["West", "Q2", 40])
    source.add_table(Table(displayName="SalesData", ref="A1:C5"))

    pivot_sheet = wb.create_sheet("Summary")
    pivot_sheet["A1"] = None
    pivot_sheet["B1"] = "Q1"
    pivot_sheet["C1"] = "Q2"
    pivot_sheet["A2"] = "East"
    pivot_sheet["B2"] = 10
    pivot_sheet["C2"] = 20
    pivot_sheet["A3"] = "West"
    pivot_sheet["B3"] = 30
    pivot_sheet["C3"] = 40
    cache = CacheDefinition(
        cacheSource=CacheSource(
            type="worksheet",
            worksheetSource=WorksheetSource(ref="A1:C5", sheet="Sales"),
        ),
        cacheFields=(
            CacheField(name="Region"),
            CacheField(name="Quarter"),
            CacheField(name="Amount"),
        ),
    )
    pivot = TableDefinition(
        name="SalesByRegion",
        cacheId=0,
        dataCaption="Values",
        location=Location(ref="A1:C3", firstHeaderRow=1, firstDataRow=2, firstDataCol=1),
    )
    pivot.cache = cache
    pivot_sheet.add_pivot(pivot)
    wb.save(path)


def test_parse_workbook_detects_pivot_table(tmp_path: Path) -> None:
    path = tmp_path / "sales_pivot.xlsx"
    _write_pivot_workbook(path)

    payload = parse_workbook(path, filename="sales_pivot.xlsx")
    by_sheet = {table["sheet"]: table for table in payload["tables"]}
    assert "Sales" in by_sheet
    assert "Summary" in by_sheet
    source = by_sheet["Sales"]
    pivot = by_sheet["Summary"]
    assert source.get("excel_table_name") == "SalesData"
    assert source.get("region_kind") == "excel_table"
    assert pivot.get("region_kind") == "pivot"
    assert pivot.get("pivot_name") == "SalesByRegion"
    assert pivot["headers"][:3] == ["column_1", "q1", "q2"]
    assert pivot["header_row"] == 1
    assert pivot["data_start_row"] == 2
    assert pivot["sample_rows"][0][:3] == ["East", 10, 20]


def test_parse_workbook_detects_table(tmp_path: Path) -> None:
    path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws.append(["Customer ID", "Company", "Email"])
    ws.append(["C1", "Acme", "a@acme.test"])
    ws.append(["C2", "Beta", "b@beta.test"])
    wb.save(path)

    payload = parse_workbook(path, filename="sample.xlsx")
    assert payload["table_count"] == 1
    table = payload["tables"][0]
    assert table["headers"] == ["customer_id", "company", "email"]
    assert table["row_count"] == 2

    profile = profile_tables(payload)
    assert profile["table_count"] == 1
    columns = profile["tables"][0]["columns"]
    assert columns[0]["likely_key"] is True

    report = interpret_tables(payload, profile, invoke=False)
    assert report["table_count"] == 1
    proposal = report["tables"][0]
    assert proposal["entity_name"]
    assert proposal["schema"]


def test_approve_table_saves_catalog_entry(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("MESHFLOW_DATA_DIR", str(tmp_path))

    from meshflow.spreadsheet.jobs import (
        approve_table,
        create_job,
        list_catalog_entries,
        load_catalog_entry,
        save_job,
        update_report_tables,
    )

    job = create_job(filename="sample.xlsx", username="poc")
    job = save_job({**job, "status": "ready"})
    table = {
        "table_id": "t0",
        "entity_name": "customers",
        "purpose": "Customer master",
        "grain": "one row per customer",
        "confidence": 0.9,
        "status": "pending_review",
        "schema": [{"name": "customer_id", "type": "string", "description": "id", "is_key": True}],
        "profiling": {"columns": []},
        "source": {"sheet": "Customers", "row_count": 2},
    }
    update_report_tables(job["job_id"], [table])
    approved = approve_table(job["job_id"], "t0", username="poc")
    assert approved["status"] == "approved"

    entries = list_catalog_entries()
    assert len(entries) == 1
    entry = load_catalog_entry(entries[0]["catalog_id"])
    assert entry is not None
    assert entry["entity_name"] == "customers"
    assert entry["filename"] == "sample.xlsx"
    assert entry["proposal"]["status"] == "approved"
    assert entry["catalog_id"] == "sample__customers"


def test_reject_table_discards_proposal(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("MESHFLOW_DATA_DIR", str(tmp_path))

    from meshflow.spreadsheet.jobs import (
        active_proposal_tables,
        create_job,
        load_report,
        reject_table,
        save_job,
        update_report_tables,
    )

    job = create_job(filename="sample.xlsx", username="poc")
    job = save_job({**job, "status": "ready"})
    keep = {
        "table_id": "t0",
        "entity_name": "customers",
        "status": "pending_review",
        "schema": [{"name": "customer_id", "type": "string"}],
        "profiling": {"columns": []},
        "source": {"sheet": "Customers", "row_count": 1},
    }
    drop = {**keep, "table_id": "t1", "entity_name": "noise"}
    update_report_tables(job["job_id"], [keep, drop])
    discarded = reject_table(job["job_id"], "t1", username="poc")
    assert discarded["status"] == "discarded"
    report = load_report(job["job_id"])
    assert report is not None
    assert len(report["tables"]) == 2
    remaining = active_proposal_tables(report["tables"])
    assert [item["table_id"] for item in remaining] == ["t0"]


def test_approve_table_materializes_silver_reference(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("MESHFLOW_DATA_DIR", str(tmp_path))

    from meshflow.storage.parquet import read_parquet_local
    from meshflow.storage.paths import prefix_path, spreadsheet_reference_silver_entity_parquet_key
    from meshflow.spreadsheet.jobs import (
        approve_table,
        create_job,
        load_catalog_entry,
        run_parse,
        store_upload,
        update_report_tables,
    )

    path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws.append(["Customer ID", "Company"])
    ws.append(["C1", "Acme"])
    ws.append(["C2", "Beta"])
    wb.save(path)

    job = create_job(filename="sample.xlsx", username="poc")
    store_upload(job["job_id"], filename="sample.xlsx", body=path.read_bytes())
    run_parse(job["job_id"])

    table = {
        "table_id": "t0",
        "entity_name": "customers",
        "purpose": "Customer master",
        "grain": "one row per customer",
        "confidence": 0.9,
        "status": "pending_review",
        "schema": [
            {"name": "customer_id", "type": "string", "description": "id", "is_key": True},
            {"name": "company", "type": "string", "description": "name"},
        ],
        "profiling": {"columns": []},
        "source": {"sheet": "Customers", "row_count": 2},
        "transformation": {
            "version": 1,
            "steps": [],
            "output_shape": {
                "entity_name": "customers",
                "schema": [
                    {"name": "customer_id", "type": "string"},
                    {"name": "company", "type": "string"},
                ],
            },
        },
    }
    update_report_tables(job["job_id"], [table])
    approve_table(job["job_id"], "t0", username="poc")

    parquet_path = prefix_path(
        tmp_path,
        spreadsheet_reference_silver_entity_parquet_key("customers"),
    )
    assert parquet_path.is_file()
    rows = read_parquet_local(parquet_path)
    assert len(rows) == 2
    assert rows[0]["customer_id"] == "C1"
    assert rows[0]["company"] == "Acme"

    entry = load_catalog_entry("sample__customers")
    assert entry is not None
    assert entry["silver_source"] == "reference"
    assert entry["silver_entity"] == "customers"
    assert entry["silver_parquet_key"] == "silver/reference/customers/data.parquet"
    assert entry["silver_row_count"] == 2


def test_compute_input_shape_hash_stable(tmp_path: Path) -> None:
    from meshflow.spreadsheet.transform import compute_input_shape

    shape_a = compute_input_shape(
        {"sheet": "Customers", "headers": ["Customer ID", "Company"]}
    )
    shape_b = compute_input_shape(
        {"sheet": "Customers", "headers": ["Customer ID", "Company"]}
    )
    assert shape_a["shape_hash"] == shape_b["shape_hash"]
    assert shape_a["column_count"] == 2


def test_apply_transformation_rename_and_cast() -> None:
    from meshflow.spreadsheet.transform import apply_transformation

    headers = ["Customer Name", "Amount"]
    rows = [["Acme", "10"], ["Beta", "20.5"]]
    spec = {
        "version": 1,
        "steps": [
            {"op": "rename_columns", "mapping": {"Customer Name": "customer_name"}},
            {"op": "cast", "columns": {"customer_name": "string", "Amount": "number"}},
        ],
        "output_shape": {
            "schema": [
                {"name": "customer_name", "type": "string"},
                {"name": "Amount", "type": "number"},
            ]
        },
    }
    out_rows, out_headers = apply_transformation(rows, headers, spec)
    assert out_headers == ["customer_name", "Amount"]
    assert out_rows[0][0] == "Acme"
    assert out_rows[1][1] == 20.5


def test_apply_transformation_keeps_rows_when_output_schema_mismatches() -> None:
    from meshflow.spreadsheet.transform import apply_transformation

    headers = ["no", "description"]
    rows = [["1896-S", "ATHENS Desk"], ["1900-S", "PARIS Chair"]]
    spec = {
        "version": 1,
        "steps": [{"op": "rename_columns", "mapping": {"no": "item_no"}}],
        "output_shape": {
            "schema": [
                {"name": "customer_id", "type": "string"},
                {"name": "company", "type": "string"},
            ]
        },
    }
    out_rows, out_headers = apply_transformation(rows, headers, spec)
    assert out_headers == ["item_no", "description"]
    assert out_rows == [["1896-S", "ATHENS Desk"], ["1900-S", "PARIS Chair"]]


def test_approve_transformation_writes_knowledge(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("MESHFLOW_DATA_DIR", str(tmp_path))

    from meshflow.spreadsheet.jobs import (
        approve_transformation,
        create_job,
        load_knowledge_entry,
        save_job,
        update_report_tables,
    )
    from meshflow.spreadsheet.transform import compute_input_shape

    job = create_job(filename="sample.xlsx", username="poc")
    job = save_job({**job, "status": "ready"})
    input_shape = compute_input_shape({"sheet": "Customers", "headers": ["customer_id", "company"]})
    transformation = {
        "version": 1,
        "steps": [{"op": "rename_columns", "mapping": {"customer_id": "customer_id"}}],
        "input_shape": input_shape,
        "output_shape": {"entity_name": "customers", "grain": "one row per customer", "schema": []},
    }
    table = {
        "table_id": "t0",
        "entity_name": "customers",
        "transformation": transformation,
        "transformation_status": "pending_review",
        "schema": [{"name": "customer_id", "type": "string"}],
    }
    update_report_tables(job["job_id"], [table])
    approve_transformation(job["job_id"], "t0", username="poc")
    kb = load_knowledge_entry("sample__customers")
    assert kb is not None
    assert kb["entity_name"] == "customers"
    assert kb["transformation"]["version"] == 1


def test_edit_transformation_resets_pending_review(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("MESHFLOW_DATA_DIR", str(tmp_path))

    from meshflow.spreadsheet.jobs import create_job, edit_transformation, load_table, save_job, update_report_tables

    job = create_job(filename="sample.xlsx")
    job = save_job({**job, "status": "ready"})
    table = {
        "table_id": "t0",
        "entity_name": "customers",
        "transformation": {"version": 1, "steps": []},
        "transformation_status": "approved",
    }
    update_report_tables(job["job_id"], [table])
    edited = edit_transformation(
        job["job_id"],
        "t0",
        {"version": 1, "steps": [{"op": "rename_columns", "mapping": {"A": "a"}}]},
    )
    assert edited["transformation_status"] == "pending_review"
    stored = load_table(job["job_id"], "t0")
    assert stored is not None
    assert stored["transformation_status"] == "pending_review"


def test_reupload_updates_last_upload_at(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("MESHFLOW_DATA_DIR", str(tmp_path))

    from meshflow.spreadsheet.jobs import (
        approve_table,
        approve_transformation,
        create_job,
        link_job_to_catalog,
        load_catalog_entry,
        record_upload_on_catalog,
        run_parse,
        save_job,
        store_upload,
        update_report_tables,
    )
    from meshflow.spreadsheet.transform import compute_input_shape

    path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws.append(["Customer ID", "Company"])
    ws.append(["C1", "Acme"])
    ws.append(["C2", "Beta"])
    wb.save(path)

    job = create_job(filename="sample.xlsx", username="poc")
    store_upload(job["job_id"], filename="sample.xlsx", body=path.read_bytes())
    run_parse(job["job_id"])

    input_shape = compute_input_shape({"sheet": "Customers", "headers": ["customer_id", "company"]})
    transformation = {
        "version": 1,
        "steps": [],
        "input_shape": input_shape,
        "output_shape": {"entity_name": "customers", "schema": []},
    }
    table = {
        "table_id": "t0",
        "entity_name": "customers",
        "status": "pending_review",
        "transformation": transformation,
        "transformation_status": "pending_review",
        "schema": [{"name": "customer_id", "type": "string"}],
    }
    update_report_tables(job["job_id"], [table])
    approve_transformation(job["job_id"], "t0", username="poc")
    approve_table(job["job_id"], "t0", username="poc")

    entry = load_catalog_entry("sample__customers")
    assert entry is not None

    job2 = create_job(filename="sample.xlsx", username="poc", linked_catalog_id="sample__customers")
    link_job_to_catalog(job2["job_id"], "sample__customers")
    record_upload_on_catalog(
        "sample__customers",
        job_id=job2["job_id"],
        uploaded_by="poc",
        input_shape_hash=input_shape["shape_hash"],
    )
    entry = load_catalog_entry("sample__customers")
    assert entry is not None
    assert entry.get("last_upload_at")
    assert entry.get("last_upload_job_id") == job2["job_id"]
    assert len(entry.get("upload_history") or []) >= 2


def test_reload_pipeline_validates_without_ai(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("MESHFLOW_DATA_DIR", str(tmp_path))

    from meshflow.spreadsheet.jobs import (
        approve_table,
        approve_transformation,
        create_job,
        load_catalog_entry,
        load_table,
        run_interpret,
        run_parse,
        run_profile,
        run_propose,
        store_upload,
        update_report_tables,
    )
    from meshflow.spreadsheet.transform import compute_input_shape

    path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws.append(["Customer ID", "Company"])
    ws.append(["C1", "Acme"])
    ws.append(["C2", "Beta"])
    wb.save(path)

    job = create_job(filename="sample.xlsx", username="poc")
    store_upload(job["job_id"], filename="sample.xlsx", body=path.read_bytes())
    run_parse(job["job_id"])

    input_shape = compute_input_shape({"sheet": "Customers", "headers": ["customer_id", "company"]})
    transformation = {
        "version": 1,
        "steps": [{"op": "rename_columns", "mapping": {}}],
        "input_shape": input_shape,
        "output_shape": {
            "entity_name": "customers",
            "grain": "one row per customer",
            "schema": [
                {"name": "customer_id", "type": "string"},
                {"name": "company", "type": "string"},
            ],
        },
    }
    table = {
        "table_id": "t0",
        "entity_name": "customers",
        "purpose": "Customers",
        "grain": "one row per customer",
        "status": "approved",
        "schema": transformation["output_shape"]["schema"],
        "transformation": transformation,
        "transformation_status": "approved",
    }
    update_report_tables(job["job_id"], [table])
    approve_transformation(job["job_id"], "t0", username="poc")
    approve_table(job["job_id"], "t0", username="poc")

    job2 = create_job(filename="sample.xlsx", username="poc", linked_catalog_id="sample__customers")
    store_upload(job2["job_id"], filename="sample.xlsx", body=path.read_bytes())
    run_parse(job2["job_id"])
    run_profile(job2["job_id"])
    run_interpret(job2["job_id"])
    run_propose(job2["job_id"])

    reloaded = load_table(job2["job_id"], "t0")
    assert reloaded is not None
    assert reloaded.get("reload_mode") is True
    assert reloaded.get("reload_validation_status") == "passed"
    assert reloaded.get("transformation_status") == "approved"


def test_validate_output_schema_detects_missing_column() -> None:
    from meshflow.spreadsheet.reload import validate_output_schema

    ok, issues = validate_output_schema(
        ["customer_id"],
        [{"name": "customer_id"}, {"name": "company"}],
    )
    assert not ok
    assert any("company" in item for item in issues)


def test_load_report_rebuilds_from_table_files(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("MESHFLOW_DATA_DIR", str(tmp_path))

    from meshflow.spreadsheet.jobs import (
        create_job,
        load_report,
        save_job,
    )
    from meshflow.storage.paths import spreadsheet_engine_job_table_key

    job = create_job(filename="sample.xlsx", username="poc")
    job = save_job({**job, "status": "ready", "table_ids": ["t0"]})
    table = {
        "table_id": "t0",
        "entity_name": "customers",
        "purpose": "Customer master",
        "grain": "one row per customer",
        "confidence": 0.9,
        "status": "pending_review",
        "schema": [{"name": "customer_id", "type": "string", "description": "id", "is_key": True}],
        "profiling": {"columns": []},
        "source": {"sheet": "Customers", "row_count": 2},
    }
    from meshflow.spreadsheet.jobs import _write_json

    _write_json(spreadsheet_engine_job_table_key(job["job_id"], "t0"), table)

    report = load_report(job["job_id"])
    assert report is not None
    assert report.get("table_count") == 1
    assert report["tables"][0]["entity_name"] == "customers"


def test_load_report_discovers_table_files_without_table_ids(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("MESHFLOW_DATA_DIR", str(tmp_path))
    monkeypatch.delenv("MESHFLOW_S3_BUCKET", raising=False)

    from meshflow.spreadsheet.jobs import _write_json, create_job, load_report, save_job
    from meshflow.storage.paths import spreadsheet_engine_job_table_key

    job = create_job(filename="sample.xlsx", username="poc")
    job = save_job({**job, "status": "ready", "table_ids": []})
    table = {
        "table_id": "t0",
        "entity_name": "customers",
        "schema": [{"name": "customer_id", "type": "string"}],
    }
    _write_json(spreadsheet_engine_job_table_key(job["job_id"], "t0"), table)

    report = load_report(job["job_id"])
    assert report is not None
    assert report.get("table_count") == 1


def _build_grouped_price_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["List Price Sheet as of 07/01/27"])
    ws.append(["Phone: +1 425 555 0100", None, None, None, None, None, None, None, None, None, None, None, None, "Page", 1])
    ws.append([None] * 7)
    ws.append(["CRONUS International Ltd."])
    ws.append([None] * 7)
    ws.append(["All Customers"])
    ws.append([None] * 7)
    ws.append([None] * 7)
    ws.append(
        [
            "No.",
            "Description",
            "Variant Code",
            "Minimum Quantity",
            None,
            "Unit of Measure Code",
            "Unit Price",
            "Starting Date",
            None,
            "Ending Date",
        ]
    )
    ws.append(["1896-S", "ATHENS Desk", None, "", None, None, None, None, None, None])
    ws.append(["", None, None, None, None, "PCS", 1000.8, "", None, None])
    ws.append(["1900-S", "PARIS Guest Chair, black", None, "", None, None, None, None, None, None])
    ws.append(["", None, None, None, None, "PCS", 192.8, "", None, None])
    wb.save(path)


def test_table_pipeline_stage_transitions() -> None:
    from meshflow.spreadsheet.stages import table_pipeline_stage

    assert table_pipeline_stage({"clean_goal": {"rows": [[1]]}, "clean_shape_status": "pending_review"}) == "clean_review"
    assert (
        table_pipeline_stage(
            {
                "clean_goal": {"rows": [[1]]},
                "clean_shape_status": "approved",
                "transformation": {"steps": [{"op": "cast"}]},
                "transformation_status": "pending_review",
            }
        )
        == "transform_review"
    )
    assert (
        table_pipeline_stage(
            {
                "clean_shape_status": "approved",
                "transformation": {"steps": [{"op": "cast"}]},
                "transformation_status": "approved",
            }
        )
        == "transform_approved"
    )
    assert table_pipeline_stage({"status": "approved"}) == "approved"


def test_apply_transformation_group_rows() -> None:
    from meshflow.spreadsheet.transform import apply_transformation

    headers = ["no", "description", "unit_of_measure_code", "unit_price"]
    rows = [
        ["1896-S", "ATHENS Desk", None, None],
        ["", None, "PCS", 1000.8],
        ["1900-S", "PARIS Guest Chair, black", None, None],
        ["", None, "PCS", 192.8],
    ]
    spec = {
        "version": 1,
        "steps": [
            {
                "op": "group_rows",
                "key_column": "no",
                "carry_columns": ["description"],
                "coalesce_columns": ["unit_of_measure_code", "unit_price"],
            },
            {"op": "filter_rows", "expr": "no != null"},
            {"op": "cast", "columns": {"unit_price": "number"}},
        ],
    }
    out_rows, out_headers = apply_transformation(rows, headers, spec)
    assert len(out_rows) == 2
    assert out_rows[0] == ["1896-S", "ATHENS Desk", "PCS", 1000.8]
    assert out_rows[1][3] == 192.8


def test_extract_table_sample_respects_byte_budget(tmp_path: Path) -> None:
    from meshflow.spreadsheet.sample import extract_table_sample

    path = tmp_path / "large.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["id", "value"])
    for idx in range(500):
        ws.append([f"row-{idx}", "x" * 200])
    wb.save(path)

    payload = parse_workbook(path)
    table = payload["tables"][0]
    sample = extract_table_sample(
        path,
        sheet=table["sheet"],
        data_start_row=table["data_start_row"],
        data_end_row=table["data_end_row"],
        min_col=table["min_col"],
        max_col=table["max_col"],
        headers=table["headers"],
        max_bytes=4096,
    )
    assert sample["sample_row_count"] < 500
    assert sample["sample_bytes"] <= 4096
    assert sample["truncated"] is True


def test_select_oracle_windows_spreads_across_sheet() -> None:
    from meshflow.spreadsheet.sample import select_oracle_windows

    rows = [[idx, f"value-{idx}"] for idx in range(2000)]
    windows = select_oracle_windows(rows, max_bytes=50_000, window_rows=100)
    assert len(windows) >= 2
    assert windows[0]["start"] == 0
    assert windows[-1]["start"] >= 1900


def test_needs_structural_cleaning_for_grouped_price_sheet(tmp_path: Path) -> None:
    from meshflow.spreadsheet.synthesize import needs_structural_cleaning

    path = tmp_path / "price_list.xlsx"
    _build_grouped_price_workbook(path)
    payload = parse_workbook(path)
    profile = profile_tables(payload)["tables"][0]
    assert needs_structural_cleaning(profile) is True


def test_induce_transformation_heuristic_for_grouped_price_sheet(tmp_path: Path) -> None:
    from meshflow.spreadsheet.sample import extract_table_sample
    from meshflow.spreadsheet.synthesize import induce_transformation_from_sample

    path = tmp_path / "price_list.xlsx"
    _build_grouped_price_workbook(path)
    payload = parse_workbook(path)
    table = payload["tables"][0]
    sample = extract_table_sample(
        path,
        sheet=table["sheet"],
        data_start_row=table["data_start_row"],
        data_end_row=table["data_end_row"],
        min_col=table["min_col"],
        max_col=table["max_col"],
        headers=table["headers"],
        header_col_offsets=list(table.get("header_col_offsets") or []),
    )
    induced = induce_transformation_from_sample(
        headers=table["headers"],
        rows=sample["rows"],
        table={"entity_name": "price_list", "schema": []},
        invoke=False,
    )
    assert induced is not None
    steps = induced["transformation"]["steps"]
    assert any(step.get("op") == "group_rows" for step in steps)
    verification = induced["induction"]["verification"]
    assert verification["passed"] is True
    assert verification["actual_row_count"] == 2


def test_propose_uses_induced_transform_for_grouped_price_sheet(tmp_path: Path) -> None:
    from meshflow.spreadsheet.propose import propose_transforms
    from meshflow.spreadsheet.sample import extract_table_sample

    path = tmp_path / "price_list.xlsx"
    _build_grouped_price_workbook(path)
    payload = parse_workbook(path)
    profile_payload = profile_tables(payload)
    report = interpret_tables(payload, profile_payload, invoke=False)
    table = payload["tables"][0]
    sample = extract_table_sample(
        path,
        sheet=table["sheet"],
        data_start_row=table["data_start_row"],
        data_end_row=table["data_end_row"],
        min_col=table["min_col"],
        max_col=table["max_col"],
        headers=table["headers"],
        header_col_offsets=list(table.get("header_col_offsets") or []),
    )
    report = propose_transforms(
        payload,
        profile_payload,
        report,
        table_samples={"t0": sample},
        invoke=False,
    )
    proposed = report["tables"][0]
    assert proposed["clean_shape_status"] == "pending_review"
    goal = proposed["clean_goal"]
    assert goal["headers"]
    assert len(goal["rows"]) == 2
    assert proposed["transformation_status"] == "awaiting_shape"
    assert not (proposed.get("transformation") or {}).get("steps")


def test_approve_clean_shape_synthesizes_transform(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("MESHFLOW_DATA_DIR", str(tmp_path))
    monkeypatch.delenv("MESHFLOW_S3_BUCKET", raising=False)

    from meshflow.spreadsheet.jobs import (
        _write_json,
        approve_clean_shape,
        create_job,
        load_table,
        run_parse,
        run_profile,
        store_upload,
    )
    from meshflow.spreadsheet.propose import propose_transforms
    from meshflow.spreadsheet.sample import extract_table_sample
    from meshflow.storage.paths import (
        spreadsheet_engine_job_parse_key,
        spreadsheet_engine_job_profile_key,
        spreadsheet_engine_job_report_key,
        spreadsheet_engine_job_table_key,
    )
    from meshflow.spreadsheet.jobs import _read_json

    path = tmp_path / "price_list.xlsx"
    _build_grouped_price_workbook(path)
    job = create_job(filename="price_list.xlsx", username="poc")
    store_upload(job["job_id"], filename="price_list.xlsx", body=path.read_bytes())
    run_parse(job["job_id"])
    run_profile(job["job_id"])

    parse_payload = _read_json(spreadsheet_engine_job_parse_key(job["job_id"]))
    profile_payload = _read_json(spreadsheet_engine_job_profile_key(job["job_id"]))
    report = interpret_tables(parse_payload, profile_payload, invoke=False)
    table = parse_payload["tables"][0]
    sample = extract_table_sample(
        path,
        sheet=table["sheet"],
        data_start_row=table["data_start_row"],
        data_end_row=table["data_end_row"],
        min_col=table["min_col"],
        max_col=table["max_col"],
        headers=table["headers"],
        header_col_offsets=list(table.get("header_col_offsets") or []),
    )
    report = propose_transforms(
        parse_payload,
        profile_payload,
        report,
        table_samples={"t0": sample},
        invoke=False,
    )
    _write_json(spreadsheet_engine_job_report_key(job["job_id"]), {**report, "job_id": job["job_id"]})
    for item in report["tables"]:
        _write_json(spreadsheet_engine_job_table_key(job["job_id"], item["table_id"]), item)

    before = load_table(job["job_id"], "t0")
    assert before["clean_shape_status"] == "pending_review"
    assert len(before["clean_goal"]["rows"]) == 2

    approve_clean_shape(job["job_id"], "t0", username="poc")
    after = load_table(job["job_id"], "t0")
    assert after["clean_shape_status"] == "approved"
    assert after["clean_goal"].get("final") is True
    assert after["pipeline_stage"] == "transform_review"
    steps = after["transformation"]["steps"]
    assert any(step.get("op") == "group_rows" for step in steps)
    assert after["transformation_status"] == "pending_review"


def test_induce_falls_back_when_oracle_is_identity(tmp_path: Path) -> None:
    """Bedrock sometimes echoes the ragged input; prefer the group_rows heuristic clean goal."""
    import json

    from meshflow.spreadsheet.propose import propose_transforms
    from meshflow.spreadsheet.sample import extract_table_sample

    path = tmp_path / "price_list.xlsx"
    _build_grouped_price_workbook(path)
    payload = parse_workbook(path)
    profile_payload = profile_tables(payload)
    table = payload["tables"][0]
    sample = extract_table_sample(
        path,
        sheet=table["sheet"],
        data_start_row=table["data_start_row"],
        data_end_row=table["data_end_row"],
        min_col=table["min_col"],
        max_col=table["max_col"],
        headers=table["headers"],
        header_col_offsets=list(table.get("header_col_offsets") or []),
    )

    def fake_invoke(system: str, user: str) -> str:
        body = json.loads(user)
        if "windows" in body:
            headers = body.get("headers") or []
            rows: list[list[object]] = []
            for window in body.get("windows") or []:
                rows.extend(window.get("rows") or [])
            return json.dumps(
                {
                    "target_headers": headers,
                    "target_rows": rows,
                    "grain": "one row per spreadsheet line",
                    "notes": ["noop"],
                }
            )
        if "output_headers" in body:
            return json.dumps(
                {
                    "steps": [{"op": "cast", "columns": {"unit_price": "number"}}],
                    "confidence": 0.9,
                    "notes": ["ai cast only"],
                }
            )
        return json.dumps(
            {
                "tables": [
                    {
                        "table_id": "t0",
                        "transformation": {"version": 1, "steps": []},
                        "transformation_confidence": 0.4,
                        "transformation_notes": ["llm empty"],
                    }
                ]
            }
        )

    report = interpret_tables(payload, profile_payload, invoke=False)
    report = propose_transforms(
        payload,
        profile_payload,
        report,
        table_samples={"t0": sample},
        invoke=fake_invoke,
    )
    proposed = report["tables"][0]
    assert proposed["clean_shape_status"] == "pending_review"
    goal = proposed["clean_goal"]
    assert goal.get("source") == "heuristic"
    assert len(goal["rows"]) == 2
    assert any("collapse" in str(n).lower() or "heuristic" in str(n).lower() for n in proposed.get("clean_shape_notes") or [])
    assert not (proposed.get("transformation") or {}).get("steps")
    assert proposed["transformation_status"] == "awaiting_shape"
