"""Detect table regions in Excel workbooks."""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

JOB_KIND = "spreadsheet_engine_parse"
MAX_SAMPLE_ROWS = 25
MIN_DATA_ROWS = 2
MIN_DATA_COLS = 2
MIN_HEADER_LABELS = 2
MIN_HEADER_DENSITY = 0.4

_REPORT_PREAMBLE_RE = re.compile(
    r"(?:\bphone\b|\bfax\b|\bpage\b|\bas of\b|\bprinted\b|\bgenerated\b|"
    r"\breport\b|\bconfidential\b|https?://|@)",
    re.IGNORECASE,
)
_LONG_DATETIME_RE = re.compile(
    r"\b(?:monday|tuesday|wednesday|thursday|friday|saturday|sunday)\b.*\b\d{4}\b",
    re.IGNORECASE,
)


def _cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _row_values(sheet: Worksheet, row_idx: int, *, min_col: int, max_col: int) -> list[Any]:
    return [
        _cell_value(sheet.cell(row=row_idx, column=col_idx).value)
        for col_idx in range(min_col, max_col + 1)
    ]


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _row_is_empty(values: list[Any]) -> bool:
    return all(_is_blank(v) for v in values)


def _looks_like_column_header(value: Any) -> bool:
    if _is_blank(value):
        return False
    if isinstance(value, (int, float, bool, datetime, date)):
        return False
    text = str(value).strip()
    if not text or len(text) > 60:
        return False
    if text.endswith(":"):
        return False
    if _REPORT_PREAMBLE_RE.search(text):
        return False
    if _LONG_DATETIME_RE.search(text):
        return False
    if len(text) > 24 and text.count(" ") >= 3:
        return False
    return True


def _header_label_indices(values: list[Any]) -> list[int]:
    return [idx for idx, value in enumerate(values) if _looks_like_column_header(value)]


def _header_band_stats(values: list[Any]) -> tuple[int, int, float]:
    label_indices = _header_label_indices(values)
    if not label_indices:
        return 0, 0, 0.0
    span = label_indices[-1] - label_indices[0] + 1
    return len(label_indices), span, len(label_indices) / span


def _data_supports_headers(
    header_values: list[Any],
    data_rows: list[list[Any]],
    *,
    col_indices: list[int],
) -> bool:
    label_cols = [idx for idx in col_indices if _looks_like_column_header(header_values[idx])]
    if len(label_cols) < MIN_HEADER_LABELS:
        return False
    if not data_rows:
        return True

    supported = 0
    for idx in label_cols:
        if any(idx < len(row) and not _is_blank(row[idx]) for row in data_rows):
            supported += 1
    return supported >= MIN_HEADER_LABELS


def _is_likely_table_header(
    header_values: list[Any],
    data_rows: list[list[Any]],
    *,
    col_indices: list[int],
) -> bool:
    label_count, span, density = _header_band_stats(header_values)
    if label_count < MIN_HEADER_LABELS:
        return False
    if span > 3 and density < MIN_HEADER_DENSITY:
        return False
    return _data_supports_headers(header_values, data_rows, col_indices=col_indices)


def _normalize_header(value: Any, *, index: int) -> str:
    if value is None:
        return f"column_{index + 1}"
    text = str(value).strip()
    if not text:
        return f"column_{index + 1}"
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_").lower()
    return slug or f"column_{index + 1}"


def _sheet_bounds(sheet: Worksheet) -> tuple[int, int, int, int]:
    if sheet.max_row is None or sheet.max_column is None:
        calculate = getattr(sheet, "calculate_dimension", None)
        if callable(calculate):
            calculate(force=True)
    min_row = sheet.min_row or 1
    max_row = sheet.max_row or 1
    min_col = sheet.min_column or 1
    max_col = sheet.max_column or 1
    return min_row, max_row, min_col, max_col


def _detect_regions(sheet: Worksheet) -> list[dict[str, Any]]:
    min_row, max_row, min_col, max_col = _sheet_bounds(sheet)
    if max_row < min_row or max_col < min_col:
        return []

    regions: list[dict[str, Any]] = []
    row = min_row
    while row <= max_row:
        header_values = _row_values(sheet, row, min_col=min_col, max_col=max_col)
        if _row_is_empty(header_values):
            row += 1
            continue

        data_start = row + 1
        data_end = row
        blank_streak = 0
        for data_row in range(data_start, max_row + 1):
            values = _row_values(sheet, data_row, min_col=min_col, max_col=max_col)
            if _row_is_empty(values):
                blank_streak += 1
                if blank_streak >= 2:
                    break
                continue
            blank_streak = 0
            data_end = data_row

        data_rows = data_end - row
        if data_rows < MIN_DATA_ROWS:
            row += 1
            continue

        non_empty_cols = [
            idx
            for idx, header in enumerate(header_values)
            if not _is_blank(header)
            or any(
                not _is_blank(
                    _row_values(sheet, r, min_col=min_col, max_col=max_col)[idx]
                )
                for r in range(data_start, data_end + 1)
            )
        ]
        if len(non_empty_cols) < MIN_DATA_COLS:
            row += 1
            continue

        col_start = min_col + non_empty_cols[0]
        col_end = min_col + non_empty_cols[-1]
        sample_rows: list[list[Any]] = []
        for sample_row in range(data_start, min(data_end + 1, data_start + MAX_SAMPLE_ROWS)):
            row_vals = _row_values(sheet, sample_row, min_col=col_start, max_col=col_end)
            if _row_is_empty(row_vals):
                continue
            sample_rows.append(row_vals)

        if not _is_likely_table_header(
            header_values,
            sample_rows,
            col_indices=non_empty_cols,
        ):
            row += 1
            continue

        headers = [
            _normalize_header(header_values[idx], index=idx) for idx in non_empty_cols
        ]
        header_col_offsets = [idx - non_empty_cols[0] for idx in non_empty_cols]
        aligned_sample_rows: list[list[Any]] = []
        for sample_row in sample_rows:
            aligned_sample_rows.append(
                [
                    sample_row[offset] if 0 <= offset < len(sample_row) else None
                    for offset in header_col_offsets
                ]
            )

        regions.append(
            {
                "sheet": sheet.title,
                "header_row": row,
                "data_start_row": data_start,
                "data_end_row": data_end,
                "min_col": col_start,
                "max_col": col_end,
                "row_count": data_rows,
                "column_count": len(headers),
                "headers": headers,
                "header_col_offsets": header_col_offsets,
                "sample_rows": aligned_sample_rows,
            }
        )
        row = data_end + 1

    return regions


def parse_workbook(path: str | Path, *, filename: str = "") -> dict[str, Any]:
    """Parse an Excel workbook into table candidates."""
    workbook_path = Path(path)
    # Random cell access in _detect_regions is incompatible with read_only mode
    # (dimensions are unset and per-cell lookups are very slow).
    workbook = load_workbook(workbook_path, data_only=True)
    tables: list[dict[str, Any]] = []
    table_index = 0
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for region in _detect_regions(sheet):
            table_id = f"t{table_index}"
            table_index += 1
            tables.append(
                {
                    "table_id": table_id,
                    "sheet": region["sheet"],
                    "header_row": region["header_row"],
                    "data_start_row": region["data_start_row"],
                    "data_end_row": region["data_end_row"],
                    "min_col": region["min_col"],
                    "max_col": region["max_col"],
                    "row_count": region["row_count"],
                    "column_count": region["column_count"],
                    "headers": region["headers"],
                    "header_col_offsets": region["header_col_offsets"],
                    "sample_rows": region["sample_rows"],
                }
            )
    workbook.close()
    return {
        "kind": JOB_KIND,
        "filename": filename or workbook_path.name,
        "sheet_count": len(workbook.sheetnames),
        "table_count": len(tables),
        "tables": tables,
    }
